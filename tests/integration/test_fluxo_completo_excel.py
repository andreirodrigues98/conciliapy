from decimal import Decimal
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from app.application.services.servico_conciliacao_planilhas import ServicoConciliacaoPlanilhas
from app.domain.enums.tipo_entrada import TipoEntrada
from app.domain.models.configuracao_conciliacao import ConfiguracaoConciliacao
from app.domain.services.calculador_resumo_conciliacao import CalculadorResumoConciliacao
from app.infrastructure.readers.leitor_planilha_excel import LeitorPlanilhaExcel
from app.infrastructure.reports.gerador_relatorio_excel import GeradorRelatorioExcel
from app.infrastructure.transformers.conversor_dataframe_registros import ConversorDataFrameRegistros
from app.infrastructure.transformers.conversor_resultados_dataframe import ConversorResultadosDataFrame


def test_fluxo_completo_gera_relatorio_excel(tmp_path: Path) -> None:
    caminho_vendas = tmp_path / "vendas.xlsx"
    caminho_pagamentos = tmp_path / "pagamentos.xlsx"
    caminho_relatorio = tmp_path / "saida" / "relatorio.xlsx"

    dataframe_vendas = pd.DataFrame(
        [
            {
                "Identificador": "PED-101",
                "Cliente": "Ana Silva",
                "Data": "05/01/2026",
                "Valor Previsto": 300.00,
            },
            {
                "Identificador": "PED-102",
                "Cliente": "Carlos Souza",
                "Data": "06/01/2026",
                "Valor Previsto": 200.00,
            },
        ]
    )

    dataframe_pagamentos = pd.DataFrame(
        [
            {
                "ID Venda": "PED-101",
                "Comprador": "Ana Silva",
                "Data Pagamento": "10/01/2026",
                "Valor Recebido": 100.00,
            },
            {
                "ID Venda": "PED-101",
                "Comprador": "Ana Silva",
                "Data Pagamento": "15/01/2026",
                "Valor Recebido": 200.00,
            },
            {
                "ID Venda": "PED-103",
                "Comprador": "Marina Lima",
                "Data Pagamento": "20/01/2026",
                "Valor Recebido": 500.00,
            },
        ]
    )

    dataframe_vendas.to_excel(
        caminho_vendas,
        index=False,
        sheet_name="Vendas",
    )

    dataframe_pagamentos.to_excel(
        caminho_pagamentos,
        index=False,
        sheet_name="Pagamentos",
    )

    configuracao = ConfiguracaoConciliacao(
        nome="Teste fluxo completo",
        tipo_entrada=TipoEntrada.DUAS_PLANILHAS,
        chave_conciliacao=("identificador",),
        tolerancia=Decimal("0.05"),
        mapeamento_vendas={
            "identificador": "Identificador",
            "cliente": "Cliente",
            "data": "Data",
            "valor_previsto": "Valor Previsto",
        },
        mapeamento_pagamentos={
            "identificador": "ID Venda",
            "cliente": "Comprador",
            "data": "Data Pagamento",
            "valor_pago": "Valor Recebido",
        },
    )

    servico = ServicoConciliacaoPlanilhas(
        leitor=LeitorPlanilhaExcel(),
        conversor=ConversorDataFrameRegistros(),
        calculador_resumo=CalculadorResumoConciliacao(),
        conversor_resultados=ConversorResultadosDataFrame(),
    )

    execucao = servico.executar(
        configuracao=configuracao,
        caminho_vendas=caminho_vendas,
        caminho_pagamentos=caminho_pagamentos,
        aba_vendas="Vendas",
        aba_pagamentos="Pagamentos",
    )

    assert len(execucao.resultados) == 3
    assert execucao.resumo.quantidade_grupos == 3
    assert execucao.resumo.quantidade_conciliados == 1
    assert execucao.resumo.total_previsto == Decimal("500.00")
    assert execucao.resumo.total_pago == Decimal("800.00")

    gerador = GeradorRelatorioExcel()

    caminho_gerado = gerador.gerar(execucao=execucao, caminho_saida=caminho_relatorio)

    assert caminho_gerado.exists()
    workbook = load_workbook(caminho_gerado)
    

    assert "Resumo" in workbook.sheetnames
    assert "Resultados" in workbook.sheetnames
    
    aba_resultados = workbook["Resultados"]

    assert aba_resultados.max_row == 4
    assert aba_resultados["A1"].value == "Chave"

    workbook.close()