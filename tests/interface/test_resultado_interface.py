import pandas as pd

from decimal import Decimal
from io import BytesIO
from openpyxl import load_workbook

from app.application.models.resultado_execucao_conciliacao import ResultadoExecucaoConciliacao
from app.domain.enums.status_conciliacao import StatusConciliacao
from app.domain.models.resumo_conciliacao import ResumoConciliacao
from app.interface.streamlit_app import formatar_moeda, gerar_relatorio_excel_bytes


def test_formatar_moeda() -> None:

    resultado = formatar_moeda(Decimal("1234.56"))

    assert resultado == "R$ 1.234,56"

def test_gerar_relatorio_excel_bytes() -> None:

    resumo = ResumoConciliacao(
        quantidade_grupos=1,
        quantidade_conciliados=1,
        total_previsto=Decimal("100.00"),
        total_pago=Decimal("100.00"),
        percentual_conciliado=Decimal("100.00"),
        contagem_por_status={StatusConciliacao.CONCILIADO: 1},
    )

    dataframe = pd.DataFrame(
        [
            {
                "Chave": "PED-001",
                "Status": (
                    StatusConciliacao
                    .CONCILIADO
                    .value
                ),
                "Total Previsto": 100.00,
                "Total Pago": 100.00,
                "Diferença": 0.00,
                "Quantidade de Previsões": 1,
                "Quantidade de Pagamentos": 1,
                "Mensagem": "Valores conciliados.",
            }
        ]
    )

    execucao = ResultadoExecucaoConciliacao(resultados=[], resumo=resumo,  dataframe_resultados=dataframe)

    conteudo = gerar_relatorio_excel_bytes(execucao=execucao)

    assert isinstance(conteudo, bytes)
    assert len(conteudo) > 0
    
    workbook = load_workbook(BytesIO(conteudo))

    assert "Resumo" in workbook.sheetnames
    assert "Resultados" in workbook.sheetnames

    workbook.close()
