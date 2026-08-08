from decimal import Decimal
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.application.models.resultado_execucao_conciliacao import ResultadoExecucaoConciliacao
from app.domain.enums.status_conciliacao import StatusConciliacao
from app.domain.models.resumo_conciliacao import ResumoConciliacao
from app.infrastructure.reports.gerador_relatorio_excel import GeradorRelatorioExcel


def criar_execucao_teste() -> ResultadoExecucaoConciliacao:

    resumo = ResumoConciliacao(
        quantidade_grupos=3,
        quantidade_conciliados=1,
        total_previsto=Decimal("500.00"),
        total_pago=Decimal("800.00"),
        percentual_conciliado=Decimal("33.33"),
        contagem_por_status={
            StatusConciliacao.CONCILIADO: 1,
            StatusConciliacao.VENDA_SEM_PAGAMENTO: 1,
            StatusConciliacao.PAGAMENTO_SEM_VENDA: 1,
        },
    )

    dataframe_resultados = pd.DataFrame(
        [
            {
                "Chave": "PED-101",
                "Status": (
                    StatusConciliacao.CONCILIADO.value
                ),
                "Total Previsto": 300.00,
                "Total Pago": 300.00,
                "Diferença": 0.00,
                "Quantidade de Previsões": 1,
                "Quantidade de Pagamentos": 2,
                "Mensagem": "Valores conciliados.",
            },
            {
                "Chave": "PED-102",
                "Status": (
                    StatusConciliacao.VENDA_SEM_PAGAMENTO.value
                ),
                "Total Previsto": 200.00,
                "Total Pago": 0.00,
                "Diferença": -200.00,
                "Quantidade de Previsões": 1,
                "Quantidade de Pagamentos": 0,
                "Mensagem": "Venda sem pagamento.",
            },
            {
                "Chave": "PED-103",
                "Status": (
                    StatusConciliacao.PAGAMENTO_SEM_VENDA.value
                ),
                "Total Previsto": 0.00,
                "Total Pago": 500.00,
                "Diferença": 500.00,
                "Quantidade de Previsões": 0,
                "Quantidade de Pagamentos": 1,
                "Mensagem": "Pagamento sem venda.",
            },
        ]
    )

    return ResultadoExecucaoConciliacao(
        resultados=[],
        resumo=resumo,
        dataframe_resultados=dataframe_resultados,
    )


def test_gerar_cria_arquivo_com_abas_esperadas(tmp_path: Path) -> None:
    execucao = criar_execucao_teste()

    gerador = GeradorRelatorioExcel()

    caminho_relatorio = (tmp_path / "relatorio.xlsx")
    caminho_gerado = gerador.gerar(execucao=execucao, caminho_saida=caminho_relatorio)

    assert caminho_gerado == caminho_relatorio
    assert caminho_gerado.exists()

    workbook = load_workbook(caminho_gerado)

    assert "Resumo" in workbook.sheetnames
    assert "Resultados" in workbook.sheetnames

    workbook.close()


def test_gerar_escreve_resumo_e_resultados(tmp_path: Path) -> None:
    execucao = criar_execucao_teste()

    gerador = GeradorRelatorioExcel()

    caminho_relatorio = (tmp_path / "relatorio.xlsx")

    gerador.gerar(execucao=execucao, caminho_saida=caminho_relatorio)

    workbook = load_workbook(caminho_relatorio)

    aba_resumo = workbook["Resumo"]

    assert (aba_resumo["A1"].value == "Resumo da Conciliação")
    assert (aba_resumo["A2"].value == "Grupos analisados")
    assert aba_resumo["B2"].value == 3

    aba_resultados = workbook["Resultados"]

    assert aba_resultados["A1"].value == "Chave"
    assert aba_resultados["B1"].value == "Status"
    assert (aba_resultados["A2"].value == "PED-101")
    assert (aba_resultados["B2"].value == StatusConciliacao.CONCILIADO.value)

    workbook.close()


def test_gerar_rejeita_extensao_diferente_de_xlsx( tmp_path: Path) -> None:
    execucao = criar_execucao_teste()

    gerador = GeradorRelatorioExcel()

    caminho_invalido = (tmp_path/ "relatorio.csv")

    with pytest.raises(ValueError, match="\\.xlsx"):
        gerador.gerar(execucao=execucao, caminho_saida=caminho_invalido,)